"""Lightweight preview window for PDF, image and Excel files.

Opened from the FilePicker popup via the "Preview" button. Renders the file in
a non-modal window positioned next to the popup so the user can verify the
download before organising it.

- Images  -> Pillow, scaled to fit, scrollable.
- PDF     -> PyMuPDF (fitz), rendered page-by-page with Prev/Next navigation.
- Excel   -> openpyxl (.xlsx/.xlsm) or xlrd (.xls), shown as a table with a
             sheet selector.
- Other   -> friendly "no preview" message.
"""

from __future__ import annotations

import threading
from pathlib import Path
from typing import Any, Callable, List, Optional, Tuple

# tkinter is only needed for the GUI window; guard the import so the pure
# helpers below stay importable on headless machines.
try:
    import tkinter as tk
    from tkinter import ttk
except ImportError:  # pragma: no cover - headless / non-GUI environments
    tk = None
    ttk = None

try:
    import customtkinter as ctk
except ImportError:  # pragma: no cover - headless / non-GUI environments
    ctk = None

from version import VERSION

# Muted dark palette (kept local so this module stays decoupled from popup.py).
_BG = "#15151d"
_BG_SECONDARY = "#1f1f2b"
_BG_FIELD = "#262633"
_ACCENT = "#5b8cff"
_ACCENT_HOVER = "#3f6fe0"
_TEXT = "#f2f2f7"
_TEXT_MUTED = "#b6b6c9"

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff", ".tif"}
PDF_EXTS = {".pdf"}
EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}

_MAX_IMAGE_DIM = 1600          # source cap; display size is zoom-controlled
_MAX_EXCEL_ROWS = 500

# PDF rendering: base DPI for the "fit" view (fast text load); zooming
# re-renders at a higher DPI so text stays crisp when enlarged.
_BASE_DPI = 90
_ZOOM_MIN = 0.5
_ZOOM_MAX = 4.0
_ZOOM_STEP = 0.25


# ----------------------------------------------------------------------
# Pure helpers (no GUI dependencies)
# ----------------------------------------------------------------------
def classify_ext(path) -> Optional[str]:
    """Return ``"image"``, ``"pdf"``, ``"excel"`` or ``None`` for a path."""
    ext = Path(path).suffix.lower()
    if ext in IMAGE_EXTS:
        return "image"
    if ext in PDF_EXTS:
        return "pdf"
    if ext in EXCEL_EXTS:
        return "excel"
    return None


def open_image(path) -> "Image.Image":
    """Open and downscale an image with Pillow (raises on failure).

    The bytes are read into memory first, so the source file is never held
    open (it stays deletable while the preview is visible on Windows).
    """
    from PIL import Image
    import io

    data = Path(path).read_bytes()
    img = Image.open(io.BytesIO(data))
    img.thumbnail((_MAX_IMAGE_DIM, _MAX_IMAGE_DIM))
    return img


def open_pdf(path) -> "pymupdf.Document":
    """Open a PDF with PyMuPDF from an in-memory buffer (raises on failure).

    Reading into memory means the source file is never locked while the
    preview is open on Windows.
    """
    try:
        import pymupdf  # PyMuPDF >= 1.24 (modern name)
    except ImportError:
        import fitz as pymupdf  # older PyMuPDF releases
    data = Path(path).read_bytes()
    return pymupdf.open(stream=data, filetype="pdf")


def read_excel(path) -> Tuple[List[str], Callable[[str], List[list]], Any]:
    """Return ``(sheet_names, loader, closeable)`` for an Excel file.

    ``loader(name)`` returns up to ``_MAX_EXCEL_ROWS`` rows as lists of raw
    cell values. ``closeable`` is the workbook/book object whose ``close()``
    releases the underlying file handle. Supports .xlsx/.xlsm via openpyxl and
    legacy .xls via xlrd. The file is read into memory first so the source is
    never held open while the preview is visible.
    """
    import io

    ext = Path(path).suffix.lower()
    data = Path(path).read_bytes()
    if ext == ".xls":
        import xlrd

        book = xlrd.open_workbook(file_contents=data)

        def load(name: str) -> List[list]:
            sh = book.sheet_by_name(name)
            rows = []
            for r in range(min(sh.nrows, _MAX_EXCEL_ROWS)):
                rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
            return rows

        return book.sheet_names(), load, book
    else:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)

        def load(name: str) -> List[list]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= _MAX_EXCEL_ROWS:
                    break
                rows.append(list(row))
            return rows

        return wb.sheetnames, load, wb


# ----------------------------------------------------------------------
# GUI preview (window or embedded pane)
# ----------------------------------------------------------------------
class PreviewWindow:
    """Previews a file, either in its own Toplevel or embedded in a container.

    When ``container`` is provided (a CTkFrame), the preview is built into it
    so it can sit inside another window (e.g. the FilePicker popup's right
    pane). Otherwise a standalone non-modal Toplevel is created.
    """

    def __init__(self, parent, file_path, container=None) -> None:
        self.file_path = Path(file_path)
        self.parent = parent

        if container is not None:
            # Embedded mode: build into the supplied frame. All preview
            # widgets pack into self.window, which is now this frame (frames
            # support pack/configure/after/winfo_* just like windows).
            self.window = container
            self.window.configure(fg_color=_BG)
            self._embedded = True
        else:
            # Standalone mode: create a non-modal Toplevel next to the parent.
            self.window = ctk.CTkToplevel(parent)
            self.window.title(
                f"FilePicker v{VERSION} — Preview: {self.file_path.name}"
            )
            self.window.configure(fg_color=_BG)
            self.window.transient(parent)
            self.window.attributes("-topmost", True)
            self.window.protocol("WM_DELETE_WINDOW", self.window.destroy)

            # Position the preview window next to the parent popup.
            parent.update_idletasks()
            x = parent.winfo_rootx() + parent.winfo_width() + 8
            y = parent.winfo_rooty()
            self.window.geometry(f"820x640+{x}+{y}")
            self._embedded = False

        kind = classify_ext(self.file_path)
        if not self.file_path.exists():
            self._show_error("File not found (it may have been moved or deleted).")
        elif kind == "image":
            self._build_image_view()
        elif kind == "pdf":
            self._build_pdf_view()
        elif kind == "excel":
            self._build_excel_view()
        else:
            self._show_error("No preview available for this file type.")

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------
    def destroy(self) -> None:
        """Release the file handle(s) and remove the preview UI."""
        self._release_resources()
        try:
            self.window.destroy()
        except tk.TclError:
            pass

    def _release_resources(self) -> None:
        """Close any open file handles so the source file can be deleted on
        Windows (an open PyMuPDF/Pillow/openpyxl handle keeps the file locked).
        """
        # PDF document: close under the render lock so an in-flight prefetch
        # thread can't touch the document while it is being closed.
        doc = getattr(self, "_doc", None)
        if doc is not None:
            lock = getattr(self, "_render_lock", None)
            try:
                if lock is not None:
                    with lock:
                        doc.close()
                else:
                    doc.close()
            except Exception:
                pass
        # Image
        img = getattr(self, "_img", None)
        if img is not None:
            try:
                img.close()
            except Exception:
                pass
        # Excel workbook
        wb = getattr(self, "_wb", None)
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    def _show_error(self, message: str) -> None:
        ctk.CTkLabel(
            self.window, text=message, text_color=_TEXT_MUTED,
            font=ctk.CTkFont(size=14), wraplength=600,
        ).pack(expand=True)

    def _make_scroll_canvas(self) -> tk.Canvas:
        # Container holds canvas + both scrollbars so vertical, horizontal and
        # pan all work. Alt+Scroll → horizontal, Shift+drag → pan (scan).
        container = tk.Frame(self.window, bg=_BG)
        container.pack(fill="both", expand=True)
        canvas = tk.Canvas(container, bg=_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)
        # Vertical scroll (plain wheel)
        canvas.bind(
            "<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"),
        )
        # Horizontal scroll via Alt+Wheel (Windows) and Alt+Button-4/5 (Linux)
        canvas.bind(
            "<Alt-MouseWheel>",
            lambda e: canvas.xview_scroll(int(-e.delta / 120), "units"),
        )
        canvas.bind("<Alt-Button-4>", lambda e: canvas.xview_scroll(-1, "units"))
        canvas.bind("<Alt-Button-5>", lambda e: canvas.xview_scroll(1, "units"))
        # Pan / move tool: Shift + click-drag (hold Shift, drag mouse)
        def _pan_start(e):
            canvas.scan_mark(e.x, e.y)
            try:
                canvas.config(cursor="fleur")
            except tk.TclError:
                pass

        def _pan_drag(e):
            canvas.scan_dragto(e.x, e.y, gain=1)

        def _pan_end(_e=None):
            try:
                canvas.config(cursor="")
            except tk.TclError:
                pass

        canvas.bind("<Shift-ButtonPress-1>", _pan_start)
        canvas.bind("<Shift-B1-Motion>", _pan_drag)
        canvas.bind("<Shift-ButtonRelease-1>", _pan_end)
        # Also allow plain middle-mouse drag as an alternative hand tool
        canvas.bind("<ButtonPress-2>", _pan_start)
        canvas.bind("<B2-Motion>", _pan_drag)
        canvas.bind("<ButtonRelease-2>", _pan_end)
        return canvas

    # ------------------------------------------------------------------
    # Images
    # ------------------------------------------------------------------
    def _build_image_view(self) -> None:
        try:
            self._img = open_image(self.file_path)
        except ImportError:
            self._show_error("Pillow is not installed. Run: pip install Pillow")
            return
        except Exception as exc:
            self._show_error(f"Could not open image: {exc}")
            return

        self._zoom = 1.0
        toolbar = ctk.CTkFrame(self.window, fg_color=_BG_SECONDARY)
        toolbar.pack(fill="x")
        self._add_zoom_controls(toolbar)
        self._canvas = self._make_scroll_canvas()
        self._canvas.bind("<Control-MouseWheel>", self._on_ctrl_wheel)
        self._render_image()
        # Fit the image to the window once it is laid out.
        self.window.after(60, self._fit_width)

    def _render_image(self) -> None:
        from PIL import Image, ImageTk
        import io

        w = max(int(self._img.width * self._zoom), 1)
        h = max(int(self._img.height * self._zoom), 1)
        img = self._img.resize((w, h), Image.LANCZOS)
        # Compress via PNG bytes: far lighter in memory than raw RGB.
        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        buf.seek(0)
        self._photo = ImageTk.PhotoImage(Image.open(buf))
        self._canvas.delete("all")
        self._canvas.create_image(0, 0, anchor="nw", image=self._photo)
        self._canvas.configure(scrollregion=(0, 0, w, h))
        self._zoom_label.configure(text=f"{int(self._zoom * 100)}%")

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------
    def _build_pdf_view(self) -> None:
        try:
            self._doc = open_pdf(self.file_path)
        except ImportError:
            self._show_error("PyMuPDF is not installed. Run: pip install PyMuPDF")
            return
        except Exception as exc:
            self._show_error(f"Could not open PDF: {exc}")
            return
        if len(self._doc) == 0:
            self._show_error("This PDF has no pages.")
            return

        self._page_index = 0
        self._zoom = 1.0
        # Page cache: page_index -> (zoom, PhotoImage). Neighbours are
        # pre-rendered in background threads so flipping pages is instant.
        self._page_cache = {}
        self._pending = set()  # page indices currently rendering in background
        # PyMuPDF documents are NOT thread-safe: serialise all page rendering
        # (main thread + prefetch threads) behind a lock.
        self._render_lock = threading.Lock()

        toolbar = ctk.CTkFrame(self.window, fg_color=_BG_SECONDARY)
        toolbar.pack(fill="x")
        self._prev_btn = ctk.CTkButton(
            toolbar, text="◀ Prev", width=80, height=30,
            fg_color=_BG_FIELD, hover_color="#33334a", text_color=_TEXT,
            command=self._prev_page,
        )
        self._prev_btn.pack(side="left", padx=(8, 4), pady=8)
        self._page_label = ctk.CTkLabel(toolbar, text="", text_color=_TEXT)
        self._page_label.pack(side="left", padx=4)
        self._next_btn = ctk.CTkButton(
            toolbar, text="Next ▶", width=80, height=30,
            fg_color=_ACCENT, hover_color=_ACCENT_HOVER, text_color="#ffffff",
            command=self._next_page,
        )
        self._next_btn.pack(side="left", padx=4, pady=8)

        self._add_zoom_controls(toolbar)

        self._canvas = self._make_scroll_canvas()
        self._canvas.bind("<Control-MouseWheel>", self._on_ctrl_wheel)
        self._render_page()
        # Fit the first page once the window is laid out.
        self.window.after(60, self._fit_width)

    def _render_page(self) -> None:
        """Render and display the current page synchronously."""
        from PIL import ImageTk

        img = self._render_page_image(self._page_index, self._zoom)
        self._photo = ImageTk.PhotoImage(img)
        self._page_cache[self._page_index] = (self._zoom, self._photo)
        self._display_photo(self._photo)
        self._update_page_controls()
        self._prefetch()

    def _render_page_image(self, idx: int, zoom: float):
        """Render one page to a compressed PIL image (no GUI needed)."""
        from PIL import Image
        import io

        with self._render_lock:  # PyMuPDF is not thread-safe
            page = self._doc.load_page(idx)
            dpi = int(_BASE_DPI * zoom)
            pix = page.get_pixmap(dpi=dpi)
            # PNG is lossless and compresses text pages extremely well, so
            # pages load fast and stay light in memory even for 50+ page docs.
            return Image.open(io.BytesIO(pix.tobytes("png")))

    def _display_photo(self, photo) -> None:
        self._canvas.delete("all")
        self._canvas.create_image(0, 0, anchor="nw", image=photo)
        self._canvas.configure(scrollregion=(0, 0, photo.width(), photo.height()))

    def _update_page_controls(self) -> None:
        self._page_label.configure(
            text=f"Page {self._page_index + 1} / {len(self._doc)}"
        )
        self._zoom_label.configure(text=f"{int(self._zoom * 100)}%")
        self._prev_btn.configure(
            state="normal" if self._page_index > 0 else "disabled"
        )
        self._next_btn.configure(
            state="normal" if self._page_index < len(self._doc) - 1 else "disabled"
        )

    def _prefetch(self) -> None:
        """Pre-render the previous and next pages in background threads.

        When the user flips to the next page, the page after it is rendered
        automatically (lookahead), so flipping stays fast.
        """
        for idx in (self._page_index - 1, self._page_index + 1):
            if (0 <= idx < len(self._doc)
                    and idx not in self._page_cache
                    and idx not in self._pending):
                self._render_page_async(idx)
        self._evict_cache()

    def _render_page_async(self, idx: int) -> None:
        self._pending.add(idx)
        zoom = self._zoom

        def work() -> None:
            try:
                img = self._render_page_image(idx, zoom)
            except Exception:
                img = None
            try:
                self.window.after(0, lambda: self._finish_async(idx, zoom, img))
            except tk.TclError:
                pass  # window closed while rendering

        threading.Thread(
            target=work, name="filepicker-pdf-render", daemon=True
        ).start()

    def _finish_async(self, idx: int, zoom: float, img) -> None:
        self._pending.discard(idx)
        if img is None or zoom != self._zoom:
            return  # stale (zoom changed while rendering)
        from PIL import ImageTk

        self._page_cache[idx] = (zoom, ImageTk.PhotoImage(img))
        self._evict_cache()

    def _evict_cache(self) -> None:
        # Keep only the current page and its immediate neighbours.
        keep = {self._page_index - 1, self._page_index, self._page_index + 1}
        for idx in [i for i in self._page_cache if i not in keep]:
            del self._page_cache[idx]

    def _prev_page(self) -> None:
        if self._page_index > 0:
            self._page_index -= 1
            self._show_cached_or_render()

    def _next_page(self) -> None:
        if self._page_index < len(self._doc) - 1:
            self._page_index += 1
            self._show_cached_or_render()

    def _show_cached_or_render(self) -> None:
        cached = self._page_cache.get(self._page_index)
        if cached is not None and cached[0] == self._zoom:
            self._photo = cached[1]
            self._display_photo(self._photo)
            self._update_page_controls()
            self._prefetch()
        else:
            self._render_page()

    # ------------------------------------------------------------------
    # Zoom helpers (shared by image & PDF views)
    # ------------------------------------------------------------------
    def _add_zoom_controls(self, toolbar) -> None:
        ctk.CTkFrame(toolbar, width=24, fg_color="transparent").pack(side="left")
        self._zoom_out_btn = ctk.CTkButton(
            toolbar, text="-", width=36, height=30,
            fg_color=_BG_FIELD, hover_color="#33334a", text_color=_TEXT,
            command=self._zoom_out,
        )
        self._zoom_out_btn.pack(side="left", padx=4, pady=8)
        self._zoom_label = ctk.CTkLabel(
            toolbar, text="100%", text_color=_TEXT, width=52,
        )
        self._zoom_label.pack(side="left", padx=4)
        self._zoom_in_btn = ctk.CTkButton(
            toolbar, text="+", width=36, height=30,
            fg_color=_ACCENT, hover_color=_ACCENT_HOVER, text_color="#ffffff",
            command=self._zoom_in,
        )
        self._zoom_in_btn.pack(side="left", padx=4, pady=8)
        self._fit_btn = ctk.CTkButton(
            toolbar, text="Fit Width", width=84, height=30,
            fg_color=_BG_FIELD, hover_color="#33334a", text_color=_TEXT,
            command=self._fit_width,
        )
        self._fit_btn.pack(side="left", padx=4, pady=8)

    def _zoom_in(self) -> None:
        self._zoom = min(self._zoom + _ZOOM_STEP, _ZOOM_MAX)
        self._invalidate_cache()
        self._render_current()

    def _zoom_out(self) -> None:
        self._zoom = max(self._zoom - _ZOOM_STEP, _ZOOM_MIN)
        self._invalidate_cache()
        self._render_current()

    def _invalidate_cache(self) -> None:
        if hasattr(self, "_page_cache"):
            self._page_cache.clear()

    def _fit_width(self) -> None:
        self.window.update_idletasks()
        canvas_w = self._canvas.winfo_width() or (self.window.winfo_width() - 20)
        if canvas_w < 100:
            canvas_w = 800
        natural_w = self._natural_width()
        if natural_w:
            self._zoom = max(min(canvas_w / natural_w, _ZOOM_MAX), _ZOOM_MIN)
        self._invalidate_cache()
        self._render_current()

    def _render_current(self) -> None:
        if hasattr(self, "_img"):
            self._render_image()
        else:
            self._render_page()

    def _natural_width(self) -> int:
        if hasattr(self, "_img"):
            return self._img.width
        page = self._doc.load_page(self._page_index)
        return page.get_pixmap(dpi=_BASE_DPI).width

    def _on_ctrl_wheel(self, event) -> None:
        if event.delta > 0:
            self._zoom_in()
        else:
            self._zoom_out()

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def _build_excel_view(self) -> None:
        try:
            self._sheets, self._load_sheet_data, self._wb = read_excel(self.file_path)
        except Exception as exc:
            self._show_error(f"Could not open spreadsheet: {exc}")
            return

        toolbar = ctk.CTkFrame(self.window, fg_color=_BG_SECONDARY)
        toolbar.pack(fill="x")
        ctk.CTkLabel(
            toolbar, text="Sheet:", text_color=_TEXT_MUTED
        ).pack(side="left", padx=(10, 4), pady=8)
        self._sheet_var = tk.StringVar(value=self._sheets[0])
        self._sheet_menu = ctk.CTkOptionMenu(
            toolbar, values=self._sheets, variable=self._sheet_var,
            command=self._load_sheet, fg_color=_BG_FIELD, button_color=_ACCENT,
            button_hover_color=_ACCENT_HOVER, width=200,
        )
        self._sheet_menu.pack(side="left", padx=4, pady=8)

        # Table
        frame = tk.Frame(self.window, bg=_BG)
        frame.pack(fill="both", expand=True)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview", background=_BG_FIELD, fieldbackground=_BG_FIELD,
            foreground=_TEXT, rowheight=24, borderwidth=0,
        )
        style.configure(
            "Treeview.Heading", background=_BG_SECONDARY, foreground=_TEXT,
            relief="flat",
        )
        style.map(
            "Treeview",
            background=[("selected", _ACCENT)],
            foreground=[("selected", "#ffffff")],
        )

        self._tree = ttk.Treeview(frame, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self._load_sheet(self._sheets[0])

    def _load_sheet(self, sheet_name: str) -> None:
        data = self._load_sheet_data(sheet_name)
        max_cols = max((len(r) for r in data), default=0)
        cols = [chr(65 + i) if i < 26 else f"Col{i + 1}" for i in range(max_cols)]

        self._tree.delete(*self._tree.get_children())
        self._tree["columns"] = cols
        for c in cols:
            self._tree.heading(c, text=c)
            self._tree.column(c, width=110, minwidth=60, stretch=False)
        for row in data:
            self._tree.insert(
                "", "end", values=[("" if v is None else str(v)) for v in row]
            )